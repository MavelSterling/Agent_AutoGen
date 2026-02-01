"""Genera archivo Excel con estimación de costos y, opcionalmente, precios por scraping/recurso."""

from __future__ import annotations

from pathlib import Path
from typing import Any, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from src.core.schemas import CostEstimate


def cost_estimate_to_excel(
    cost: CostEstimate,
    path: Path,
    scraped_rows: List[dict[str, Any]] | None = None,
    resources: Optional[List[str]] = None,
) -> None:
    """Escribe la estimación de costos en un archivo .xlsx.

    - Hoja 'Estimacion': rangos, drivers y supuestos de volumen.
    - Hoja 'Precios_nube': precios obtenidos por scraping (provider, servicio, unidad, precio, region).
    - Hoja 'Costos_por_recurso': un fila por recurso del proyecto con precio estimado (si hay resources y scraped_rows).
    """
    wb = Workbook()
    ws_est = wb.active
    ws_est.title = "Estimacion"

    # Encabezado
    ws_est["A1"] = "Concepto"
    ws_est["B1"] = "Valor"
    ws_est["A1"].font = ws_est["B1"].font = Font(bold=True)

    row = 2
    ws_est.cell(row=row, column=1, value="Rango bajo")
    ws_est.cell(row=row, column=2, value=cost.range_low)
    row += 1
    ws_est.cell(row=row, column=1, value="Rango medio")
    ws_est.cell(row=row, column=2, value=cost.range_mid)
    row += 1
    ws_est.cell(row=row, column=1, value="Rango alto")
    ws_est.cell(row=row, column=2, value=cost.range_high)
    row += 1
    ws_est.cell(row=row, column=1, value="Drivers")
    ws_est.cell(row=row, column=2, value=", ".join(cost.drivers))
    row += 1
    ws_est.cell(row=row, column=1, value="Supuestos de volumen")
    ws_est.cell(row=row, column=2, value=", ".join(cost.volume_assumptions))

    if scraped_rows:
        ws_scraped = wb.create_sheet("Precios_nube")
        headers = ["provider", "servicio", "unidad", "precio", "region", "fuente"]
        for col, h in enumerate(headers, start=1):
            ws_scraped.cell(row=1, column=col, value=h if h != "fuente" else "Fuente")
            ws_scraped.cell(row=1, column=col).font = Font(bold=True)
        for r, item in enumerate(scraped_rows, start=2):
            for c, key in enumerate(headers, start=1):
                val = item.get(key, "")
                ws_scraped.cell(row=r, column=c, value=val)

    if resources and scraped_rows:
        ws_recurso = wb.create_sheet("Costos_por_recurso")
        rec_headers = ["recurso", "servicio_equivalente", "unidad", "precio_estimado", "region", "fuente"]
        for col, h in enumerate(rec_headers, start=1):
            ws_recurso.cell(row=1, column=col, value=h if h != "fuente" else "Fuente")
            ws_recurso.cell(row=1, column=col).font = Font(bold=True)
        for r, recurso in enumerate(resources, start=2):
            ws_recurso.cell(row=r, column=1, value=recurso)
            match = _match_resource_to_price(recurso, scraped_rows)
            if match:
                ws_recurso.cell(row=r, column=2, value=match.get("servicio", ""))
                ws_recurso.cell(row=r, column=3, value=match.get("unidad", ""))
                ws_recurso.cell(row=r, column=4, value=match.get("precio", ""))
                ws_recurso.cell(row=r, column=5, value=match.get("region", ""))
                ws_recurso.cell(row=r, column=6, value=match.get("fuente", ""))
            else:
                ws_recurso.cell(row=r, column=2, value="consultar")
                ws_recurso.cell(row=r, column=4, value="consultar")
                ws_recurso.cell(row=r, column=6, value="")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _match_resource_to_price(recurso: str, scraped_rows: List[dict[str, Any]]) -> dict[str, Any] | None:
    """Busca en scraped_rows un servicio que coincida con el nombre del recurso (parcial o mapeo conocido)."""
    recurso_lower = recurso.lower()
    # Mapeo conocido: nombre en requirements -> subcadena en servicio de precios
    alias_map = {
        "container registry": ["container registry", "acr", "registry"],
        "container instance": ["container instance", "aci", "container instances"],
        "storage account": ["storage", "blob", "storage account"],
        "cosmos db": ["cosmos db", "cosmos"],
        "key vault": ["key vault", "keyvault"],
        "application insights": ["application insights", "app insights"],
        "log analytics": ["log analytics", "logs"],
        "monitor": ["monitor", "azure monitor"],
        "service bus": ["service bus", "servicebus"],
        "event grid": ["event grid"],
        "event hubs": ["event hubs", "event hub"],
        "api management": ["api management", "apim"],
        "azure ai search": ["search", "cognitive search", "ai search"],
        "container apps": ["container apps", "container app"],
        "azure web apps": ["web app", "app service", "web apps"],
        "azure bot services": ["bot", "bot service"],
        "azure ai foundry": ["ai foundry", "foundry", "openai"],
    }
    keywords = alias_map.get(recurso_lower, [recurso_lower])
    for row in scraped_rows:
        servicio = (row.get("servicio") or "").lower()
        if any(kw in servicio for kw in keywords):
            return row
    for row in scraped_rows:
        if recurso_lower in (row.get("servicio") or "").lower():
            return row
    return None
